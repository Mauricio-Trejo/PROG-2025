import csv
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Carpeta de origen de los CSV y destino del Excel
CARPETA_DATOS = r"D:\PIA_ProgBas\PIA_PrograBas\Guardar_datos"
ARCHIVO_EXCEL = os.path.join(CARPETA_DATOS, "calidad_aire.xlsx")

# Estilos
ESTILO_ENCABEZADO = {
    "font": Font(bold=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="4F81BD"),
    "alignment": Alignment(horizontal="center", vertical="center")
}
ESTILO_CELDA = Alignment(horizontal="center", vertical="center")

def exportar_a_excel(ciudad):
    ciudad = ciudad.strip().lower()
    archivo_csv = os.path.join(CARPETA_DATOS, f"datos_{ciudad}.csv")
    
    if not os.path.exists(archivo_csv):
        print(f"Archivo CSV no encontrado para la ciudad '{ciudad}'.")
        return

    # Cargar o crear libro de Excel
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Eliminar hoja previa si existe
    if ciudad.capitalize() in wb.sheetnames:
        del wb[ciudad.capitalize()]
        
    ws = wb.create_sheet(title=ciudad.capitalize())

    with open(archivo_csv, newline='', encoding='utf-8') as f:
        lector = csv.reader(f)
        for i, fila in enumerate(lector):
            for j, celda in enumerate(fila):
                celda_excel = ws.cell(row=i+1, column=j+1, value=celda)
                if i == 0:
                    celda_excel.font = ESTILO_ENCABEZADO["font"]
                    celda_excel.fill = ESTILO_ENCABEZADO["fill"]
                    celda_excel.alignment = ESTILO_ENCABEZADO["alignment"]
                else:
                    celda_excel.alignment = ESTILO_CELDA

    # Ajustar el ancho de columna automáticamente
    for col in ws.columns:
        max_length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(ARCHIVO_EXCEL)
    print(f"Datos exportados a Excel: {ARCHIVO_EXCEL}")

# Ejecución directa
if __name__ == "__main__":
    ciudad = input("Ingrese el nombre de la ciudad para exportar a Excel: ")
    exportar_a_excel(ciudad)
